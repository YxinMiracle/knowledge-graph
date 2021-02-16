import openpyxl
from py2neo import Graph, Node, NodeMatcher
from utils.tools import get_md5
from elasticsearch import Elasticsearch
from elasticsearch import helpers

es = Elasticsearch(['127.0.0.1:9200'])


def get_data():
    """
    测试函数
    :return:
    """
    wb = openpyxl.load_workbook("c语言考纲.xlsx")
    sh = wb["Sheet1"]
    rows_list = list(sh.rows)[1:]
    for row in rows_list:
        print(row[0].value, row[1].value, row[2].value)
    wb.close()


def create_node(wb, label):
    g = Graph("http://localhost:7474", username="neo4j", password="newxin.001206")
    sh = wb["Sheet1"]
    rows_list = list(sh.rows)[1:]
    # 这个循环时将所有的节点加入到neo4j图数据库中
    for row in rows_list:
        matcher = NodeMatcher(g)
        name1 = row[0].value
        name2 = row[1].value
        if not matcher.match(label, name=name1).first():
            node = g.create(Node(label, name=name1, _id=get_md5(name1)))
            print(node)
        if not matcher.match(label, name=name2).first():
            node = g.create(Node(label, name=name2, _id=get_md5(name2)))
            print(node)

    # 这个是将Node建立起关系
    for row in rows_list:
        name1 = row[0].value
        name2 = row[1].value
        relationship_name = row[2].value
        _name1_md5 = get_md5(name1)
        _name2_md5 = get_md5(name2)
        query = "match(p:`%s`),(q:`%s`) where p._id='%s' and q._id='%s' create (p)-[r:%s]->(q)" % (
            label, label, _name1_md5, _name2_md5, relationship_name)
        g.run(query)
        print("yes!")


# 更新节点的属性，对专有名词的解释 ------------------------------------------------有问题
def update_node(wb):
    g = Graph("http://localhost:7474", username="neo4j", password="newxin.001206")
    sh = wb["Sheet1"]
    rows_list = list(sh.rows)[1:]
    actions = []
    for row in rows_list:
        n = row[0].value  # 专有名词
        try:
            exp = row[1].value.replace("\n", " ")  # 对应的解释
            query = "match(p:`%s`) where p.name='%s' set p.desc='%s' return p" % (label, n, exp)
            ret = g.run(query)
            print(n, exp, ret)
        except:
            # 含有特殊字符，目前不能解决
            print("含有特殊字符，不能进行插入")
        exp = row[1].value.replace("\n", " ")
        d = {
            "_index": "n_example",
            "_type": "doc",
            "_source": {
                "n": n,
                "example": exp
            }
        }
        actions.append(d)
    helpers.bulk(es, actions=actions)


def set_question():
    wb = openpyxl.load_workbook("c语言题目.xlsx")
    label = "C语言"
    sh = wb["Sheet1"]
    rows_list = list(sh.rows)[1:]
    actions = []
    for row in rows_list:
        question = row[0].value
        A = row[1].value
        B = row[2].value
        C = row[3].value
        D = row[4].value
        ans = row[5].value
        tags = " ".join((row[6].value).split("/"))
        print(question, A, B, C, D, ans, tags)
        d = {
            "_index": "question",
            "_type": "doc",
            "_source": {
                "question": question,
                "A":A,
                "B":B,
                "C":C,
                "D":D,
                "ans":ans,
                "tags":tags,
                "label":label
            }
        }
        actions.append(d)
    helpers.bulk(es, actions=actions)


# 得到所有的tags的name，并且加入到es的tags索引中
def get_tags_name():
    g = Graph("http://localhost:7474", username="neo4j", password="newxin.001206")
    query = "match (n) return n"
    nodes = g.run(query).data()
    actions = []
    for node in nodes:
        node_name = node.get('n').get("name")
        print(node_name)
        d = {
            "_index": "tags",
            "_type": "doc",
            "_source": {
                "title": node_name
            }
        }
        actions.append(d)
    helpers.bulk(es, actions=actions)

def get_labels():
    """将数据库中的label返回"""
    # g = Graph("http://localhost:7474", username="neo4j", password="newxin.001206")
    # query = "match (n) return n"
    # nodes = g.run(query).data()
    pass

if __name__ == '__main__':
    # wb = openpyxl.load_workbook("c#.xlsx")
    # label = "C#语言"
    # create_node(wb,label)
    # wb = openpyxl.load_workbook("c语言专有名字解释.xlsx")
    # # update_node(wb, label)
    # get_tags_name()
    # update_node(wb)
    # set_question()
    get_labels()